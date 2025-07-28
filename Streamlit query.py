import streamlit as st
import openpyxl
import os
from datetime import datetime
import cohere
import smtplib
from email.message import EmailMessage
import mimetypes
import traceback

# --- Custom CSS for white background and dark green font ---
st.markdown("""
    <style>
        body, .stApp {
            background-color: #FFFFFF !important;
            color: #08312A !important;
        }
        .stButton > button {
            color: #FFFFFF !important;
            background: #00E47C !important;
            border: none !important;
            border-radius: 8px !important;
        }
        .stTextInput > div > div > input,
        .stTextArea > div > textarea {
            background-color: #FFFFFF !important;
            color: #08312A !important;
            border: 1px solid #00E47C !important;
        }
        label,
        .css-1c7y2kd,
        .stRadio label,
        .stCheckbox label,
        [data-testid="stRadioLabel"],
        [data-testid="stSelectboxLabel"],
        [data-testid="stTextInputLabel"],
        [data-testid="stTextAreaLabel"] {
            color: #08312A !important;
        }
        h1, h2, h3, h4, h5, h6,
        .stMarkdown, .stSubheader, .stText, .stAlert {
            color: #08312A !important;
        }
        .stAlert {
            background-color: #00E47C !important;
            color: #08312A !important;
        }
        ::selection {
            background: #00E47C;
            color: #FFFFFF;
        }
        .stTextArea textarea {
            color: #08312A !important;
            background-color: #FFFFFF !important;
        }
        .stRadio span, .stCheckbox span, .stSelectbox span,
        [data-testid="stRadioItem"] > div > div > span,
        [data-testid="stSelectboxOption"] > div > span,
        [data-testid="stRadioItemLabel"] {
            color: #08312A !important;
        }
        [data-testid="stRadioItem"] *, [data-testid="stSelectboxOption"] * {
            color: #08312A !important;
        }
    </style>
""", unsafe_allow_html=True)

# --- Cohere Client ---
co = cohere.Client(api_key="nwQE8lzxJVgUHFiBSj3cVc8JBjuNwyZJrJjRgteb")

# --- Directory for Excel export ---
excel_directory = os.path.expanduser("~/Desktop/Query_Answers")
os.makedirs(excel_directory, exist_ok=True)

questions = {
    "name": {"question": "1.1. Quel est votre nom et prénom ? 😊", "options": []},
    "email": {"question": "1.2. Quelle est votre adresse e-mail ? 📧", "options": []},
    "business_unit": {
        "question": "1.3. Quelle est votre Business Unit ? 🏢",
        "options": ["Pet Vet", "Avian", "Ruminant", "Swine", "Equine", "Pet Retail"]
    },
    "supplier_name": {"question": "2.1. Quel est le nom du fournisseur ? 🏭", "options": []},
    "product_code": {
        "question": "2.2. Le produit a-t-il déjà un code existant ? 🔢",
        "options": ["Oui", "Non"]
    },
    "product_code_yes": {"question": "Veuillez indiquer le SKU actuel : 🆔", "options": []},
    "product_code_no": {"question": "Veuillez indiquer le SKU précédent ou similaire : 🆔", "options": []},
    "product_description": {
        "question": "2.3. Fournissez une brève description du produit : 📝 ou description rattachée en automatique?",
        "options": []
    },
    "supplier_conditions": {
        "question": "3.1. Le fournisseur impose-t-il une quantité minimale de commande, ou taille de lot? 📦",
        "options": ["Oui", "Non"]
    },
    "quantity_minimum_yes": {
        "question": "Indiquez la quantité minimale requise : 🔢, ou à négocier? - Y a t il des paliers de prix avec remise possible?",
        "options": []
    },
    "coverage_duration": {
        "question": "3.2. Avez-vous une idée de la durée de couverture estimée ? ⏳",
        "options": ["Oui", "Non"]
    },
    "coverage_duration_yes": {
        "question": "Indiquez la durée de couverture estimée (en mois) : 📅, selon l'historique des ventes en N-1",
        "options": []
    },
    "supplier_location": {
        "question": "4.1. Où est basé le fournisseur ? 🌍",
        "options": ["En France", "Europe", "Grand export"]
    },
    "availability_delay": {
        "question": "4.2. Quel est le délai estimé pour la mise à disposition du produit ? ⏱️",
        "options": []
    },
    "storage_location": {
        "question": "5.1. le SKU accompagne-t-il des produits finis? 📍",
        "options": ["Oui", "Non"]
    },
    "sku_open": {
        "question": "5.2. le SKU doit-il être ouvert dans Bi connect?",
        "options": ["Oui", "Non"]
    },
    "sku_frequency": {
        "question": "5.3. le SKU est-il ponctuel ou récurrent?",
        "options": []
    },
    "dotation": {
        "question": "6.1. Le produit est-il destiné à une dotation ? 🎁",
        "options": ["Oui", "Non"]
    },
    "dotation_yes": {
        "question": "Veuillez indiquer les délais impératifs de livraison sur le 3PL : 🚚",
        "options": []
    },
    "additional_requirements": {
        "question": "7.1. Y a-t-il des exigences supplémentaires ? ❓",
        "options": []
    }
    # !!! DO NOT include "final" as a question !!!
}

FINAL_KEY = "final"

def get_next_question(answer, previous_question):
    mapping = {
        "name": "email",
        "email": "business_unit",
        "business_unit": "supplier_name",
        "supplier_name": "product_code",
        "product_code": {"Oui": "product_code_yes", "Non": "product_code_no"},
        "product_code_yes": "product_description",
        "product_code_no": "product_description",
        "product_description": "supplier_conditions",
        "supplier_conditions": {"Oui": "quantity_minimum_yes", "Non": "coverage_duration"},
        "quantity_minimum_yes": "coverage_duration",
        "coverage_duration": {"Oui": "coverage_duration_yes", "Non": "supplier_location"},
        "coverage_duration_yes": "supplier_location",
        "supplier_location": "availability_delay",
        "availability_delay": "storage_location",
        "storage_location": "sku_open",
        "sku_open": "sku_frequency",
        "sku_frequency": "dotation",
        "dotation": {"Oui": "dotation_yes", "Non": "additional_requirements"},
        "dotation_yes": "additional_requirements",
        "additional_requirements": FINAL_KEY
    }
    next_question = mapping.get(previous_question)
    if isinstance(next_question, dict):
        return next_question.get(answer)
    return next_question

def save_answers_to_excel(recommendation, ai_recommendation):
    SENDER_EMAIL = "andres.osorio_garzon@boehringer-ingelheim.com"
    SENDER_PASSWORD = "Libna197169*"  # Considera usar variable de entorno en producción
    TO_EMAIL_FIXED = "andres.osorio_garzon@boehringer-ingelheim.com"
    user_name = st.session_state.user_answers.get("name")
    user_email = st.session_state.user_answers.get("email")
    if not user_name:
        st.warning("Le nom de l'utilisateur est manquant.")
        return
    current_date = datetime.now().strftime("%d-%m-%Y")
    file_name = f"{user_name}_{current_date}.xlsx"
    file_path = os.path.join(excel_directory, file_name)
    try:
        # Guardar Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Réponses"
        sheet.cell(row=1, column=1, value="Question")
        sheet.cell(row=1, column=2, value="Réponse")
        for idx, (question, answer) in enumerate(st.session_state.user_answers.items(), start=2):
            sheet.cell(row=idx, column=1, value=question)
            sheet.cell(row=idx, column=2, value=answer)
        sheet.cell(row=len(st.session_state.user_answers) + 2, column=1, value="Recommandations")
        sheet.cell(row=len(st.session_state.user_answers) + 2, column=2, value=recommendation)
        sheet.cell(row=len(st.session_state.user_answers) + 3, column=1, value="Recommandations IA")
        sheet.cell(row=len(st.session_state.user_answers) + 3, column=2, value=ai_recommendation)
        workbook.save(file_path)
        st.success(f"Les réponses ont été enregistrées dans {file_path}")

        # Construir resumen de respuestas
        summary = "Résumé des réponses:\n\n"
        for question, answer in st.session_state.user_answers.items():
            summary += f"- {question}: {answer}\n"
        summary += f"\n{recommendation}\n"

        # Asunto del correo
        subject = f"Outils Marketing - {current_date} - {user_name}"

        # Cuerpo del correo
        body = (
            f"{summary}\n"
            f"\nLe fichier Excel est adjunto. "
            f"La contraseña para abrir el archivo est: {file_name.replace('.xlsx','')}\n"
        )

        # Preparar mensaje
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = SENDER_EMAIL
        # Destinatarios: fijo + usuario encuestado
        recipients = [TO_EMAIL_FIXED]
        if user_email and user_email != TO_EMAIL_FIXED:
            recipients.append(user_email)
        msg["To"] = ";".join(recipients)
        msg.set_content(body)

        # Adjuntar archivo
        with open(file_path, "rb") as f:
            file_data = f.read()
            guess = mimetypes.guess_type(file_path)[0]
            if guess:
                maintype, subtype = guess.split("/")
            else:
                maintype, subtype = "application", "octet-stream"
            msg.add_attachment(file_data, maintype=maintype, subtype=subtype, filename=file_name)

        # Enviar correo
        print("\n4️⃣ Enviando correo...")
        try:
            with smtplib.SMTP("authsmtp.boehringer.com", 587) as smtp:
                smtp.starttls()
                smtp.login(SENDER_EMAIL, SENDER_PASSWORD)
                print("🔐 Autenticado correctamente")

                result = smtp.send_message(msg)
                if result:
                    print(f"⚠️ Algunos destinatarios fallaron: {result}")
                    st.error(f"Algunos destinatarios fallaron: {result}")
                else:
                    print("🎉 ¡CORREO ENVIADO EXITOSAMENTE!")
                    print(f"📧 Enviado a {len(recipients)} destinatarios")
                    print("⏰ El correo debería llegar en los próximos minutos")
                    st.success(f"Correo enviado a {', '.join(recipients)}")
        except Exception as e:
            print(f"❌ Error en el envío: {e}")
            print(traceback.format_exc())
            st.error(f"Error en el envío: {e}")

        print("\n" + "=" * 50)
    except PermissionError:
        st.error(f"Impossible d'enregistrer le fichier. Vérifiez les permissions pour le chemin : {file_path}")
    except Exception as e:
        st.error(f"Erreur inattendue : {e}")

def show_recommendation():
    recommendation = "Recommandations :\n"
    product_code = st.session_state.user_answers.get("product_code")
    if product_code == "Non":
        recommendation += "- Assurez-vous de créer un nouveau code dans le système avant de passer commande.\n"
    quantity_minimum = st.session_state.user_answers.get("supplier_conditions")
    if quantity_minimum == "Oui":
        recommendation += "- Recommandez une analyse de consommation historique pour ajuster les hypothèses de réapprovisionnement.\n"
    supplier_location = st.session_state.user_answers.get("supplier_location")
    if supplier_location == "Grand export":
        recommendation += "- Prévoir un délai logistique plus long et anticiper les commandes.\n"
    dotation = st.session_state.user_answers.get("dotation")
    if dotation == "Oui":
        recommendation += "- Priorisez la planification logistique avec le 3PL pour respecter les délais impératifs.\n"

    def get_ai_recommendation(answers):
        try:
            prompt = "Voici les réponses d'un utilisateur à un questionnaire :\n"
            for question, answer in answers.items():
                prompt += f"- {question}: {answer}\n"
            prompt += "Basé sur ces réponses, fournissez des recommandations supplémentaires pertinentes 30 mots max:"
            response = co.generate(prompt=prompt, model="xlarge")
            return response.generations[0].text.strip()
        except Exception as e:
            return f"Erreur lors de la génération des recommandations IA : {str(e)}"

    ai_recommendation = get_ai_recommendation(st.session_state.user_answers)
    recommendation += f"\nRecommandations IA :\n{ai_recommendation}"
    st.text_area("Recommandations", recommendation)
    if st.button("Enregistrer les réponses"):
        save_answers_to_excel(recommendation, ai_recommendation)

def main():
    st.title("Outil Marketing Survey")
    st.write("Merci de répondre aux questions pour obtenir des recommandations personnalisées.")

    # Init session state
    if "current_question" not in st.session_state:
        st.session_state.current_question = "name"
    if "user_answers" not in st.session_state:
        st.session_state.user_answers = {}

    current_question_key = st.session_state.current_question

    if current_question_key != FINAL_KEY:
        question_data = questions.get(current_question_key)
        widget_key = f"widget_{current_question_key}"
        st.subheader(question_data["question"])
        with st.form(key=f"form_{current_question_key}"):
            if question_data["options"]:
                answer = st.radio("Choisissez une option :", question_data["options"], key=widget_key)
            else:
                answer = st.text_input("Votre réponse :", key=widget_key)
            submitted = st.form_submit_button("Suivant")
            if submitted:
                if answer:
                    st.session_state.user_answers[current_question_key] = answer
                    next_question = get_next_question(answer, current_question_key)
                    if next_question:
                        st.session_state.current_question = next_question
                        # Clear widget state for next question to avoid carryover
                        next_widget_key = f"widget_{next_question}"
                        if next_widget_key in st.session_state:
                            del st.session_state[next_widget_key]
                    else:
                        st.session_state.current_question = FINAL_KEY
                else:
                    st.warning("Veuillez entrer une réponse avant de continuer.")

        if widget_key not in st.session_state or not st.session_state[widget_key]:
            st.info("Veuillez répondre avant de cliquer sur Suivant.")
    else:
        # Only show recommendations, not a question or input!
        st.header("Fin du formulaire 🏁")
        show_recommendation()

if __name__ == "__main__":
    main()
